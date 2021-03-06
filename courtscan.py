"""
What you need to get this working:

1. Python 3 
2. Pip 
3. Use Pip to install the following packages: openpyxl, beautifulsoup4, requests, selenium
4. Microsoft Excel or LibreOffice/OpenOffice
5. Google Chrome (latest version, 83)
6. ChromeDriver

After this is all ready, set up your excel file. I named mine testcourt.xlsx
Go to our web based excel sheet and copy the court case numbers into the first column

"""

import openpyxl
import requests
import time
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys

chrome_options = Options()
chrome_options.add_argument("--headless")
chrome_options.add_argument("--window-size=1920x1080")
#Change the file name to whatever you named your file!
file_name = 'testcourt.xlsx'
base_url = 'http://www.fcmcclerk.com/case/search'
second_url = 'http://www.fcmcclerk.com/case/view'
webpage = requests.get(base_url)
timeout = 10 

####open file and get ready to traverse 
court_file = openpyxl.load_workbook(file_name)
cur_sheet = court_file.active
cur_row = 1
cur_col = 1 
cur_cell = cur_sheet.cell(row = cur_row, column = cur_col)

#loop through first column which contains our court case numbers
while(cur_cell.value != None):
	driver = webdriver.Chrome(options=chrome_options, executable_path='chromedriver')
	driver.get(base_url)
	case_num = cur_cell.value
	print("Searching: " + case_num)
	
	#attempt to search our current case number
	try:
		#wait for webpage to load
		WebDriverWait(driver, timeout).until(EC.visibility_of_element_located((By.ID, "case_number")))
		
		#enter case number into search field
		case_search = driver.find_element_by_id("case_number")
		case_search.send_keys(case_num)
		#click search button to perform our search
		search_button = driver.find_element_by_id("search_button_1")
		search_button.click()
		#click view button 
		#print("waiting 3 seconds") #force script to wait seems to allow all elements to load sufficiently
		time.sleep(3)
		#print("waiting done")
		view_button = driver.find_element_by_xpath("//input[@value='View']")
		#print("Clicking view button")
		view_button.click()
		#print("View Button clicked")
		#switch to new tab opened
		driver.switch_to.window(driver.window_handles[1])
		try:
			WebDriverWait(driver, timeout).until(EC.visibility_of_element_located((By.ID, "evnt_table")))
			WebDriverWait(driver, timeout).until(EC.visibility_of_element_located((By.ID, "dsp_table")))
			disp_table = driver.find_element_by_id("dsp_table")
			event_table = driver.find_element_by_id("evnt_table")
			#event table loaded, now grab the date of hearing
			#wow holy shit this works!!
			dispo = driver.find_element_by_xpath('//*[@id="dsp_table"]/tbody/tr[2]/td[1]')
			if dispo.text == "CLOSED":
				print("CLOSED status: fetching reason")
				dispo_data = driver.find_element_by_xpath('//*[@id="dsp_table"]/tbody/tr[2]/td[3]')
				write_cell = cur_sheet.cell(row = cur_row, column = cur_col+1) 
				write_cell.value = dispo_data.text
				
			else:
				#print(dispo.text)
				print("Found event table. Fetching hearing date.")
				event_date = driver.find_element_by_xpath('//*[@id="evnt_table"]/tbody/tr[2]/td[2]')
				#store the date as text in a handy variable
				event_text_date = event_date.text
				#add the date to our spreadsheet in the right column!
				write_cell = cur_sheet.cell(row = cur_row, column = cur_col+1) 
				write_cell.value = event_text_date
				print("Attempting to fetch defendant address.")
				#check to make sure defendant is listed where we think it should be
				try:
					WebDriverWait(driver, timeout).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="pty_table"]/tbody/tr[5]/td[5]')))
					is_defendant = driver.find_element_by_xpath('//*[@id="pty_table"]/tbody/tr[5]/td[5]')
					#defendant is listed in expected place
					if is_defendant.text == "DEFENDANT": 
						def_address = driver.find_element_by_xpath('//*[@id="pty_table"]/tbody/tr[6]/td[2]')
					#if an address is listed for the defendant
						if(def_address):
							def_city = driver.find_element_by_xpath('//*[@id="pty_table"]/tbody/tr[7]/td[2]')
							def_zip = driver.find_element_by_xpath('//*[@id="pty_table"]/tbody/tr[7]/td[4]')
							write_cell = cur_sheet.cell(row = cur_row, column = cur_col+2)
							write_cell.value = def_address.text + " " + def_city.text + " " + def_zip.text 
						else:
							print("Defendant address not found.")
				except TimeoutException:
					print("Could not find address")
					court_file.save(file_name)
					driver.quit()
			
			
		except TimeoutException:
			print("Could not find event table with date")
			try:
				disp_table = driver.find_element_by_id("dsp_table")
				dispo = driver.find_element_by_xpath('//*[@id="dsp_table"]/tbody/tr[2]/td[1]')
				if dispo.text == "CLOSED":
					print("CLOSED status: fetching reason")
					dispo_data = driver.find_element_by_xpath('//*[@id="dsp_table"]/tbody/tr[2]/td[3]')
					write_cell = cur_sheet.cell(row = cur_row, column = cur_col+1) 
					write_cell.value = dispo_data.text
			except: 
				print("ERROR")
				write_cell = cur_sheet.cell(row = cur_row, column = cur_col+1) 
				write_cell.value = "ERROR"
				court_file.save(file_name)
				
			driver.quit()
		
		
	#if page doesn't load say so and close browser
	except TimeoutException:
		print("Could not load webpage")
		driver.quit()
		
	cur_row += 1
	cur_cell = cur_sheet.cell(row = cur_row, column = cur_col)
	court_file.save(file_name)
	driver.quit()
###end loop

print("Saving data")
driver.quit()
court_file.save(file_name)
court_file.close()
print("Data entry finished")

#Now all the court dates are in the appripriate column next to the associated court case number. You can copy/paste from your personal excel file into the google docs one. 
